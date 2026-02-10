"""
Budget vs Forecast Comparison Script
Extracts and compares key monthly metrics from Budget and Forecast files
"""

import pandas as pd
import openpyxl
from pathlib import Path
import warnings
warnings.filterwarnings('ignore')

# File paths
BUDGET_FILE = '/home/user/OBF/Budget consol file.xlsx'
FORECAST_FILE = '/home/user/OBF/forecast_output/Forecast_Transparency_Report.xlsx'

def explore_excel_structure(filepath, file_label):
    """
    Explore the structure of an Excel file
    """
    print(f"\n{'='*80}")
    print(f"{file_label}: {Path(filepath).name}")
    print(f"{'='*80}")

    wb = openpyxl.load_workbook(filepath, data_only=True)
    print(f"\nSheet names: {wb.sheetnames}")

    sheet_data = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n{'-'*60}")
        print(f"Sheet: {sheet_name}")
        print(f"{'-'*60}")

        # Get dimensions
        max_row = ws.max_row
        max_col = ws.max_column
        print(f"Dimensions: {max_row} rows x {max_col} columns")

        # Print first 15 rows to understand structure
        print("\nFirst 15 rows (sample):")
        rows_data = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(15, max_row), values_only=True), 1):
            # Clean the row data for display
            clean_row = [cell if cell is not None else '' for cell in row[:15]]  # First 15 columns
            rows_data.append(clean_row)
            print(f"Row {row_idx}: {clean_row}")

        sheet_data[sheet_name] = {
            'workbook': wb,
            'worksheet': ws,
            'max_row': max_row,
            'max_col': max_col,
            'sample_rows': rows_data
        }

    return sheet_data

def extract_budget_metrics(sheet_data):
    """
    Extract monthly metrics from the budget file
    """
    print(f"\n{'='*80}")
    print("EXTRACTING BUDGET METRICS")
    print(f"{'='*80}")

    budget_metrics = {}

    # Try to find the main data sheet - common names include 'Budget', 'Consolidated', 'Summary', etc.
    possible_sheets = ['Budget', 'Consol', 'Consolidated', 'Summary', 'Total', 'All']

    for sheet_name in sheet_data.keys():
        print(f"\nChecking sheet: {sheet_name}")
        ws = sheet_data[sheet_name]['worksheet']

        # Try reading as pandas DataFrame to make extraction easier
        try:
            df = pd.read_excel(BUDGET_FILE, sheet_name=sheet_name, header=None)
            print(f"DataFrame shape: {df.shape}")
            print("\nFirst 20 rows:")
            print(df.head(20))

            # Look for date/month columns and metric rows
            # Common patterns: months in columns, metrics in rows
            # OR months in rows, metrics in columns

        except Exception as e:
            print(f"Error reading sheet {sheet_name}: {e}")

    return budget_metrics

def extract_forecast_metrics(sheet_data):
    """
    Extract monthly metrics from the forecast transparency report
    """
    print(f"\n{'='*80}")
    print("EXTRACTING FORECAST METRICS")
    print(f"{'='*80}")

    forecast_metrics = {}

    for sheet_name in sheet_data.keys():
        print(f"\nChecking sheet: {sheet_name}")
        ws = sheet_data[sheet_name]['worksheet']

        # Try reading as pandas DataFrame
        try:
            df = pd.read_excel(FORECAST_FILE, sheet_name=sheet_name, header=None)
            print(f"DataFrame shape: {df.shape}")
            print("\nFirst 20 rows:")
            print(df.head(20))

        except Exception as e:
            print(f"Error reading sheet {sheet_name}: {e}")

    return forecast_metrics

def main():
    """
    Main execution function
    """
    print("BUDGET VS FORECAST COMPARISON ANALYSIS")
    print("="*80)

    # Step 1: Explore both files
    print("\n\nSTEP 1: EXPLORING FILE STRUCTURES")
    print("="*80)

    budget_sheets = explore_excel_structure(BUDGET_FILE, "BUDGET FILE")
    forecast_sheets = explore_excel_structure(FORECAST_FILE, "FORECAST FILE")

    # Step 2: Extract budget metrics
    print("\n\nSTEP 2: EXTRACTING METRICS")
    print("="*80)

    budget_metrics = extract_budget_metrics(budget_sheets)
    forecast_metrics = extract_forecast_metrics(forecast_sheets)

    # Step 3: Compare and compute variances
    print("\n\nSTEP 3: VARIANCE ANALYSIS")
    print("="*80)
    print("(To be implemented after understanding data structure)")

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\nERROR: {e}")
        import traceback
        traceback.print_exc()
