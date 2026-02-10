"""
Comprehensive Forecast vs Actuals Analysis
Compares BB Forecast to Total Book Actuals on a % of GBV basis
"""
import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook(
    "/home/user/OBF/Forecast Baseline Outputs v3.6 (new collections & impairment).xlsx",
    data_only=True
)
ws = wb["forecast vs actuals"]

# =============================================================================
# SECTION 1: Extract BB Forecast detailed data (Rows 5-44)
# Columns Q(17) onwards = Oct 2025+
# =============================================================================
print("=" * 120)
print("SECTION 1: BB FORECAST DATA (Oct 2025 - Sep 2026)")
print("=" * 120)

# Month columns for forecast: Q=17 (Oct 25) through AB=28 (Sep 26)
forecast_months = []
for col_num in range(17, 29):  # Q=17 to AB=28
    col_letter = get_column_letter(col_num)
    header = ws[f'{col_letter}1'].value
    if header:
        forecast_months.append((col_num, col_letter, str(header)[:10]))
    else:
        # Try row 4
        header = ws[f'{col_letter}4'].value
        forecast_months.append((col_num, col_letter, str(header)[:10] if header else f"Col{col_letter}"))

print(f"\nForecast months: {[m[2] for m in forecast_months]}")

# BB Forecast metrics mapping (row -> (metric, segment))
bb_forecast_rows = {
    5: ("OpeningGBV", "NON PRIME"),
    6: ("OpeningGBV", "NRP-L"),
    7: ("OpeningGBV", "NRP-M"),
    8: ("OpeningGBV", "NRP-S"),
    9: ("OpeningGBV", "PRIME"),
    10: ("Coll_Principal", "NON PRIME"),
    11: ("Coll_Principal", "NRP-L"),
    12: ("Coll_Principal", "NRP-M"),
    13: ("Coll_Principal", "NRP-S"),
    14: ("Coll_Principal", "PRIME"),
    15: ("Coll_Interest", "NON PRIME"),
    16: ("Coll_Interest", "NRP-L"),
    17: ("Coll_Interest", "NRP-M"),
    18: ("Coll_Interest", "NRP-S"),
    19: ("Coll_Interest", "PRIME"),
    20: ("ClosingGBV", "NON PRIME"),
    21: ("ClosingGBV", "NRP-L"),
    22: ("ClosingGBV", "NRP-M"),
    23: ("ClosingGBV", "NRP-S"),
    24: ("ClosingGBV", "PRIME"),
    25: ("ClosingNBV", "NON PRIME"),
    26: ("ClosingNBV", "NRP-L"),
    27: ("ClosingNBV", "NRP-M"),
    28: ("ClosingNBV", "NRP-S"),
    29: ("ClosingNBV", "PRIME"),
    30: ("InterestRevenue", "NON PRIME"),
    31: ("InterestRevenue", "NRP-L"),
    32: ("InterestRevenue", "NRP-M"),
    33: ("InterestRevenue", "NRP-S"),
    34: ("InterestRevenue", "PRIME"),
    35: ("Gross_Impairment", "NON PRIME"),
    36: ("Gross_Impairment", "NRP-L"),
    37: ("Gross_Impairment", "NRP-M"),
    38: ("Gross_Impairment", "NRP-S"),
    39: ("Gross_Impairment", "PRIME"),
    40: ("Net_Impairment", "NON PRIME"),
    41: ("Net_Impairment", "NRP-L"),
    42: ("Net_Impairment", "NRP-M"),
    43: ("Net_Impairment", "NRP-S"),
    44: ("Net_Impairment", "PRIME"),
}

# Extract BB Forecast data
bb_data = {}
for row_num, (metric, segment) in bb_forecast_rows.items():
    key = (metric, segment)
    bb_data[key] = {}
    for col_num, col_letter, month_label in forecast_months:
        val = ws[f'{col_letter}{row_num}'].value
        bb_data[key][month_label] = val if val is not None else 0.0

# Print BB Forecast summary for ALL segments combined
print("\n--- BB Forecast: Monthly Totals (All Segments) ---")
metrics_to_show = ["OpeningGBV", "Coll_Principal", "Coll_Interest", "ClosingGBV",
                   "ClosingNBV", "InterestRevenue", "Gross_Impairment", "Net_Impairment"]
segments = ["NON PRIME", "NRP-L", "NRP-M", "NRP-S", "PRIME"]

# Build totals
bb_totals = {}
for metric in metrics_to_show:
    bb_totals[metric] = {}
    for _, _, month_label in forecast_months:
        total = 0
        for seg in segments:
            val = bb_data.get((metric, seg), {}).get(month_label, 0)
            total += val if val else 0
        bb_totals[metric][month_label] = total

print(f"\n{'Metric':<25}", end="")
for _, _, ml in forecast_months:
    print(f" {ml:>14}", end="")
print()
print("-" * (25 + 15 * len(forecast_months)))

for metric in metrics_to_show:
    print(f"{metric:<25}", end="")
    for _, _, ml in forecast_months:
        val = bb_totals[metric][ml]
        print(f" {val/1e6:>13.2f}m", end="")
    print()

# =============================================================================
# SECTION 2: BB as % of Opening GBV
# =============================================================================
print("\n\n" + "=" * 120)
print("SECTION 2: BB FORECAST - KEY METRICS AS % OF OPENING GBV")
print("=" * 120)

print(f"\n{'Metric':<25}", end="")
for _, _, ml in forecast_months:
    print(f" {ml:>12}", end="")
print()
print("-" * (25 + 13 * len(forecast_months)))

for metric in ["Coll_Principal", "Coll_Interest", "InterestRevenue", "Gross_Impairment", "Net_Impairment"]:
    print(f"{metric:<25}", end="")
    for _, _, ml in forecast_months:
        ogbv = bb_totals["OpeningGBV"][ml]
        val = bb_totals[metric][ml]
        pct = (val / ogbv * 100) if ogbv != 0 else 0
        print(f" {pct:>11.2f}%", end="")
    print()

# Coverage Ratio (Provision / GBV)
print(f"{'Coverage Ratio':<25}", end="")
for _, _, ml in forecast_months:
    cgbv = bb_totals["ClosingGBV"][ml]
    cnbv = bb_totals["ClosingNBV"][ml]
    provision = cgbv - cnbv
    cr = (provision / cgbv * 100) if cgbv != 0 else 0
    print(f" {cr:>11.2f}%", end="")
print()

# =============================================================================
# SECTION 3: Extract ACTUALS (Total Book) data
# Rows 194-368, Columns H(8) to P(16) = Jan-Sep 2025
# =============================================================================
print("\n\n" + "=" * 120)
print("SECTION 3: ACTUALS (TOTAL BOOK) DATA")
print("=" * 120)

# Actuals month columns: H=8 (Jan 25) to P=16 (Sep 25), then Q=17 (Oct 25) onwards
actuals_months = []
for col_num in range(8, 29):  # H=8 through AB=28
    col_letter = get_column_letter(col_num)
    header = ws[f'{col_letter}1'].value
    if header:
        actuals_months.append((col_num, col_letter, str(header)[:10]))

print(f"\nActuals months available: {[m[2] for m in actuals_months]}")

# Key actuals rows
actuals_key_rows = {
    196: "Collections_NonPrime",
    197: "Collections_NRP_S",
    198: "Collections_NRP_M",
    199: "Collections_Prime",
    200: "Collections_Total",
    207: "Contra_NonPrime",
    208: "Contra_NRP_S",
    209: "Contra_NRP_M",
    210: "Contra_Prime",
    211: "Contra_Total",
    222: "TotalColl_incContra",
    240: "ClosingGBV",
    250: "AvgGBV",
    260: "ClosingNBV",
    270: "AvgNBV",
    280: "Revenue",
    281: "Revenue_%AvgGBV",
    291: "GrossImpairment",
    292: "GrossImp_%Revenue",
    297: "GrossImp_%AvgGBV",
    307: "RAM_exclDS",
    313: "RAM_exclDS_%AvgGBV",
    323: "DebtSaleGain",
    329: "DSGain_%AvgGBV",
    339: "NetImpairment",
    345: "NetImp_%AvgGBV",
    355: "RAM_inclDS",
    361: "RAM_inclDS_%AvgGBV",
}

# Extract actuals data
actuals_data = {}
for row_num, metric_name in actuals_key_rows.items():
    actuals_data[metric_name] = {}
    for col_num, col_letter, month_label in actuals_months:
        val = ws[f'{col_letter}{row_num}'].value
        actuals_data[metric_name][month_label] = val

# Print key actuals metrics
print("\n--- Total Book Actuals: Key Metrics (£m) ---")
actuals_abs_metrics = ["Collections_Total", "TotalColl_incContra", "ClosingGBV", "AvgGBV",
                       "ClosingNBV", "AvgNBV", "Revenue", "GrossImpairment", "NetImpairment",
                       "RAM_exclDS", "DebtSaleGain", "RAM_inclDS"]

print(f"\n{'Metric':<25}", end="")
for _, _, ml in actuals_months:
    print(f" {ml:>14}", end="")
print()
print("-" * (25 + 15 * len(actuals_months)))

for metric_name in actuals_abs_metrics:
    print(f"{metric_name:<25}", end="")
    for _, _, ml in actuals_months:
        val = actuals_data[metric_name].get(ml)
        if val is not None and isinstance(val, (int, float)):
            print(f" {val/1e6:>13.2f}m", end="")
        else:
            print(f" {'N/A':>14}", end="")
    print()

# Print % metrics
print("\n--- Total Book Actuals: % Metrics ---")
actuals_pct_metrics = ["Revenue_%AvgGBV", "GrossImp_%AvgGBV", "NetImp_%AvgGBV",
                       "RAM_exclDS_%AvgGBV", "DSGain_%AvgGBV", "RAM_inclDS_%AvgGBV"]

print(f"\n{'Metric':<25}", end="")
for _, _, ml in actuals_months:
    print(f" {ml:>12}", end="")
print()
print("-" * (25 + 13 * len(actuals_months)))

for metric_name in actuals_pct_metrics:
    print(f"{metric_name:<25}", end="")
    for _, _, ml in actuals_months:
        val = actuals_data[metric_name].get(ml)
        if val is not None and isinstance(val, (int, float)):
            print(f" {val*100:>11.2f}%", end="")
        else:
            print(f" {'N/A':>12}", end="")
    print()

# =============================================================================
# SECTION 4: COMPARISON - BB Forecast vs Total Book Actuals (% of GBV basis)
# For the overlap period (Oct 2025 onwards where both have data)
# =============================================================================
print("\n\n" + "=" * 120)
print("SECTION 4: COMPARISON - BB FORECAST vs TOTAL BOOK ACTUALS")
print("=" * 120)

# First, calculate BB forecast as % of BB Opening GBV
# Then compare to Total Book actuals as % of Total Book Avg GBV

# Overlap months (Oct 2025 onwards where actuals exist)
print("\n--- Checking which months have BOTH BB forecast AND total book actuals ---")
overlap_months = []
for _, _, ml in forecast_months:
    bb_ogbv = bb_totals["OpeningGBV"].get(ml, 0)
    act_avgbv = actuals_data["AvgGBV"].get(ml)
    has_bb = bb_ogbv > 0
    has_act = act_avgbv is not None and isinstance(act_avgbv, (int, float)) and act_avgbv > 0
    print(f"  {ml}: BB OpenGBV={bb_ogbv/1e6:.1f}m, Actuals AvgGBV={'N/A' if not has_act else f'{act_avgbv/1e6:.1f}m'} -> {'BOTH' if has_bb and has_act else 'BB only' if has_bb else 'Neither'}")
    if has_bb and has_act:
        overlap_months.append(ml)

if overlap_months:
    print(f"\nOverlap months with both datasets: {overlap_months}")

    print("\n--- KEY COMPARISON: Metrics as % of GBV ---")
    print(f"\n{'Metric':<30} {'Source':<10}", end="")
    for ml in overlap_months:
        print(f" {ml:>12}", end="")
    print()
    print("-" * (40 + 13 * len(overlap_months)))

    # Collections as % of GBV
    print(f"{'Collections % GBV':<30} {'BB':<10}", end="")
    for ml in overlap_months:
        bb_coll = bb_totals["Coll_Principal"][ml] + bb_totals["Coll_Interest"][ml]
        bb_ogbv = bb_totals["OpeningGBV"][ml]
        pct = (bb_coll / bb_ogbv * 100) if bb_ogbv != 0 else 0
        print(f" {pct:>11.2f}%", end="")
    print()
    print(f"{'':30} {'Actuals':<10}", end="")
    for ml in overlap_months:
        act_coll_raw = actuals_data["Collections_Total"].get(ml, 0)
        act_coll = act_coll_raw if isinstance(act_coll_raw, (int, float)) else 0
        act_gbv_raw = actuals_data["AvgGBV"].get(ml, 1)
        act_gbv = act_gbv_raw if isinstance(act_gbv_raw, (int, float)) else 1
        pct = (act_coll / act_gbv * 100) if act_gbv else 0
        print(f" {pct:>11.2f}%", end="")
    print()

    # Revenue as % of GBV
    print(f"\n{'Revenue % GBV':<30} {'BB':<10}", end="")
    for ml in overlap_months:
        bb_rev = bb_totals["InterestRevenue"][ml]
        bb_ogbv = bb_totals["OpeningGBV"][ml]
        pct = (bb_rev / bb_ogbv * 100) if bb_ogbv != 0 else 0
        print(f" {pct:>11.2f}%", end="")
    print()
    print(f"{'':30} {'Actuals':<10}", end="")
    for ml in overlap_months:
        val = actuals_data["Revenue_%AvgGBV"].get(ml)
        if val is not None and isinstance(val, (int, float)):
            print(f" {val*100:>11.2f}%", end="")
        else:
            print(f" {'N/A':>12}", end="")
    print()

    # Gross Impairment as % of GBV
    print(f"\n{'Gross Imp % GBV':<30} {'BB':<10}", end="")
    for ml in overlap_months:
        bb_imp = bb_totals["Gross_Impairment"][ml]
        bb_ogbv = bb_totals["OpeningGBV"][ml]
        pct = (bb_imp / bb_ogbv * 100) if bb_ogbv != 0 else 0
        print(f" {pct:>11.2f}%", end="")
    print()
    print(f"{'':30} {'Actuals':<10}", end="")
    for ml in overlap_months:
        val = actuals_data["GrossImp_%AvgGBV"].get(ml)
        if val is not None and isinstance(val, (int, float)):
            print(f" {val*100:>11.2f}%", end="")
        else:
            print(f" {'N/A':>12}", end="")
    print()

    # Net Impairment as % of GBV
    print(f"\n{'Net Imp % GBV':<30} {'BB':<10}", end="")
    for ml in overlap_months:
        bb_nimp = bb_totals["Net_Impairment"][ml]
        bb_ogbv = bb_totals["OpeningGBV"][ml]
        pct = (bb_nimp / bb_ogbv * 100) if bb_ogbv != 0 else 0
        print(f" {pct:>11.2f}%", end="")
    print()
    print(f"{'':30} {'Actuals':<10}", end="")
    for ml in overlap_months:
        val = actuals_data["NetImp_%AvgGBV"].get(ml)
        if val is not None and isinstance(val, (int, float)):
            print(f" {val*100:>11.2f}%", end="")
        else:
            print(f" {'N/A':>12}", end="")
    print()

# =============================================================================
# SECTION 5: Historical BB % of Total Book (actuals period)
# Need to get BB actuals from the model's actuals data
# =============================================================================
print("\n\n" + "=" * 120)
print("SECTION 5: HISTORICAL BB AS % OF TOTAL BOOK (Jan-Sep 2025)")
print("=" * 120)

# The BB forecast section only shows Oct 2025 onwards
# But the BB SUMMARY section (rows 49-193) may have historical BB data too
# Let me check what's in the summary section for actuals months

# Check rows 49-100 for BB summary data
print("\n--- Checking BB Summary Section (rows 49-100) for historical data ---")
for row_num in range(49, 100):
    label = ws[f'F{row_num}'].value
    if label and str(label).strip():
        # Check if there's data in actuals columns (H-P)
        has_actuals = False
        for col_num in range(8, 17):
            val = ws[f'{get_column_letter(col_num)}{row_num}'].value
            if val is not None:
                has_actuals = True
                break

        has_forecast = False
        val_q = ws[f'Q{row_num}'].value
        if val_q is not None:
            has_forecast = True

        if has_actuals or has_forecast:
            vals = []
            for col_num in [8, 16, 17, 18]:  # Jan, Sep, Oct, Nov
                val = ws[f'{get_column_letter(col_num)}{row_num}'].value
                vals.append(f"{val}" if val is not None else "None")
            print(f"  Row {row_num}: {str(label)[:50]:<50} Jan={vals[0]:<15} Sep={vals[1]:<15} Oct={vals[2]:<15} Nov={vals[3]:<15}")

# =============================================================================
# SECTION 6: Detailed segment-level BB Forecast analysis
# =============================================================================
print("\n\n" + "=" * 120)
print("SECTION 6: BB FORECAST BY SEGMENT (Oct-Dec 2025)")
print("=" * 120)

# For each segment, show key metrics for first 3 forecast months
first_3_months = forecast_months[:3]

for segment in segments:
    print(f"\n--- {segment} ---")
    ogbv_vals = [bb_data.get(("OpeningGBV", segment), {}).get(ml, 0) for _, _, ml in first_3_months]
    cgbv_vals = [bb_data.get(("ClosingGBV", segment), {}).get(ml, 0) for _, _, ml in first_3_months]
    cnbv_vals = [bb_data.get(("ClosingNBV", segment), {}).get(ml, 0) for _, _, ml in first_3_months]
    coll_p_vals = [bb_data.get(("Coll_Principal", segment), {}).get(ml, 0) for _, _, ml in first_3_months]
    coll_i_vals = [bb_data.get(("Coll_Interest", segment), {}).get(ml, 0) for _, _, ml in first_3_months]
    rev_vals = [bb_data.get(("InterestRevenue", segment), {}).get(ml, 0) for _, _, ml in first_3_months]
    gimp_vals = [bb_data.get(("Gross_Impairment", segment), {}).get(ml, 0) for _, _, ml in first_3_months]
    nimp_vals = [bb_data.get(("Net_Impairment", segment), {}).get(ml, 0) for _, _, ml in first_3_months]

    print(f"  {'Metric':<25}", end="")
    for _, _, ml in first_3_months:
        print(f" {ml:>14}", end="")
    print()

    print(f"  {'Opening GBV (£m)':<25}", end="")
    for v in ogbv_vals:
        print(f" {v/1e6:>13.2f}m", end="")
    print()

    print(f"  {'Closing GBV (£m)':<25}", end="")
    for v in cgbv_vals:
        print(f" {v/1e6:>13.2f}m", end="")
    print()

    print(f"  {'Coll_P + Coll_I (£m)':<25}", end="")
    for i in range(3):
        total_coll = (coll_p_vals[i] or 0) + (coll_i_vals[i] or 0)
        print(f" {total_coll/1e6:>13.2f}m", end="")
    print()

    print(f"  {'Collections % GBV':<25}", end="")
    for i in range(3):
        total_coll = abs((coll_p_vals[i] or 0) + (coll_i_vals[i] or 0))
        ogbv = ogbv_vals[i] if ogbv_vals[i] else 1
        print(f" {total_coll/ogbv*100:>13.2f}%", end="")
    print()

    print(f"  {'Interest Revenue (£m)':<25}", end="")
    for v in rev_vals:
        print(f" {(v or 0)/1e6:>13.2f}m", end="")
    print()

    print(f"  {'Revenue % GBV':<25}", end="")
    for i in range(3):
        ogbv = ogbv_vals[i] if ogbv_vals[i] else 1
        print(f" {(rev_vals[i] or 0)/ogbv*100:>13.2f}%", end="")
    print()

    print(f"  {'Gross Impairment (£m)':<25}", end="")
    for v in gimp_vals:
        print(f" {(v or 0)/1e6:>13.2f}m", end="")
    print()

    print(f"  {'Gross Imp % GBV':<25}", end="")
    for i in range(3):
        ogbv = ogbv_vals[i] if ogbv_vals[i] else 1
        print(f" {(gimp_vals[i] or 0)/ogbv*100:>13.2f}%", end="")
    print()

    print(f"  {'Net Impairment (£m)':<25}", end="")
    for v in nimp_vals:
        print(f" {(v or 0)/1e6:>13.2f}m", end="")
    print()

    print(f"  {'Net Imp % GBV':<25}", end="")
    for i in range(3):
        ogbv = ogbv_vals[i] if ogbv_vals[i] else 1
        print(f" {(nimp_vals[i] or 0)/ogbv*100:>13.2f}%", end="")
    print()

    # Coverage ratio
    print(f"  {'Coverage Ratio':<25}", end="")
    for i in range(3):
        if cgbv_vals[i] and cgbv_vals[i] > 0:
            provision = (cgbv_vals[i] or 0) - (cnbv_vals[i] or 0)
            print(f" {provision/cgbv_vals[i]*100:>13.2f}%", end="")
        else:
            print(f" {'N/A':>14}", end="")
    print()

    # GBV decline rate
    print(f"  {'GBV Decline Rate':<25}", end="")
    for i in range(3):
        if ogbv_vals[i] and ogbv_vals[i] > 0:
            decline = ((cgbv_vals[i] or 0) - ogbv_vals[i]) / ogbv_vals[i] * 100
            print(f" {decline:>13.2f}%", end="")
        else:
            print(f" {'N/A':>14}", end="")
    print()

# =============================================================================
# SECTION 7: Extract VARIANCES section
# =============================================================================
print("\n\n" + "=" * 120)
print("SECTION 7: PRE-COMPUTED VARIANCES (from sheet)")
print("=" * 120)

# Check what's in the variance section
print("\n--- Scanning variance section (rows 369-450) ---")
for row_num in range(369, 450):
    label = ws[f'F{row_num}'].value
    if label and str(label).strip():
        vals = []
        for col_num in [17, 18, 19]:  # Oct, Nov, Dec 2025
            val = ws[f'{get_column_letter(col_num)}{row_num}'].value
            if val is not None:
                vals.append(f"{val:.4f}" if isinstance(val, float) and abs(val) < 1 else f"{val/1e6:.2f}m" if isinstance(val, (int, float)) else str(val))
            else:
                vals.append("None")
        print(f"  Row {row_num}: {str(label)[:55]:<55} Oct={vals[0]:<18} Nov={vals[1]:<18} Dec={vals[2]:<18}")

# =============================================================================
# SECTION 8: ACTUALS Period Analysis (Jan-Sep 2025)
# Look at total book trends for context
# =============================================================================
print("\n\n" + "=" * 120)
print("SECTION 8: TOTAL BOOK ACTUALS TREND (Jan-Sep 2025)")
print("=" * 120)

act_months_hist = [(cn, cl, ml) for cn, cl, ml in actuals_months if ml < "2025-10"]
print(f"\n{'Metric':<25}", end="")
for _, _, ml in act_months_hist:
    print(f" {ml:>12}", end="")
print()
print("-" * (25 + 13 * len(act_months_hist)))

for metric_name in ["ClosingGBV", "AvgGBV", "Revenue", "GrossImpairment", "NetImpairment", "RAM_inclDS"]:
    print(f"{metric_name:<25}", end="")
    for _, _, ml in act_months_hist:
        val = actuals_data[metric_name].get(ml)
        if val is not None:
            print(f" {val/1e6:>11.2f}m", end="")
        else:
            print(f" {'N/A':>12}", end="")
    print()

print(f"\n{'% of Avg GBV':<25}", end="")
for _, _, ml in act_months_hist:
    print(f" {ml:>12}", end="")
print()
print("-" * (25 + 13 * len(act_months_hist)))

for metric_name in ["Revenue_%AvgGBV", "GrossImp_%AvgGBV", "NetImp_%AvgGBV", "RAM_inclDS_%AvgGBV"]:
    print(f"{metric_name:<25}", end="")
    for _, _, ml in act_months_hist:
        val = actuals_data[metric_name].get(ml)
        if val is not None and isinstance(val, (int, float)):
            print(f" {val*100:>11.2f}%", end="")
        else:
            print(f" {'N/A':>12}", end="")
    print()

# =============================================================================
# SECTION 9: Actuals segment-level data
# =============================================================================
print("\n\n" + "=" * 120)
print("SECTION 9: ACTUALS BY SEGMENT (checking rows 283-300)")
print("=" * 120)

# Check segment-level actuals for revenue and impairment
for row_num in range(280, 350):
    label = ws[f'F{row_num}'].value
    if label and str(label).strip():
        vals = []
        for col_num in [8, 12, 16]:  # Jan, May, Sep
            val = ws[f'{get_column_letter(col_num)}{row_num}'].value
            if val is not None:
                if isinstance(val, float):
                    if abs(val) < 10:
                        vals.append(f"{val*100:.2f}%")
                    else:
                        vals.append(f"£{val/1e6:.2f}m")
                else:
                    vals.append(str(val)[:20])
            else:
                vals.append("None")
        print(f"  Row {row_num}: {str(label)[:50]:<50} Jan={vals[0]:<15} May={vals[1]:<15} Sep={vals[2]:<15}")

# =============================================================================
# SECTION 10: BB Forecast Summary section - check for % comparisons
# =============================================================================
print("\n\n" + "=" * 120)
print("SECTION 10: BB FORECAST SUMMARY SECTION (rows 49-193) - Key % Metrics")
print("=" * 120)

for row_num in range(49, 194):
    label_e = ws[f'E{row_num}'].value
    label_f = ws[f'F{row_num}'].value
    label = label_e or label_f
    if label and str(label).strip():
        vals = []
        for col_num in [17, 18, 19, 20]:  # Oct, Nov, Dec 25, Jan 26
            val = ws[f'{get_column_letter(col_num)}{row_num}'].value
            if val is not None:
                if isinstance(val, float):
                    if abs(val) < 2:  # Likely a percentage or ratio
                        vals.append(f"{val*100:.2f}%")
                    else:
                        vals.append(f"£{val/1e6:.2f}m")
                else:
                    vals.append(str(val)[:18])
            else:
                vals.append("None")
        print(f"  Row {row_num}: {str(label)[:55]:<55} Oct={vals[0]:<15} Nov={vals[1]:<15} Dec={vals[2]:<15} Jan={vals[3]:<15}")

wb.close()
print("\n\nAnalysis complete.")
