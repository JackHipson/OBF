"""
Parse the 'forecast vs actuals' sheet from the baseline outputs file.
Print EVERY cell value to understand the exact structure.
"""

import openpyxl
from openpyxl.utils import get_column_letter

# File path
file_path = "/home/user/OBF/Forecast Baseline Outputs v3.6 (new collections & impairment).xlsx"

# Open the workbook
print("Opening workbook...")
wb = openpyxl.load_workbook(file_path, data_only=True)

# Access the sheet
sheet_name = "forecast vs actuals"
if sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"Successfully opened sheet: {sheet_name}\n")
else:
    print(f"Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}")
    exit()

# Get the max column (up to BF = column 58)
max_col = min(ws.max_column, 58)  # BF is the 58th column
max_row = 100

print(f"Sheet dimensions: {ws.max_row} rows x {ws.max_column} columns")
print(f"Will scan: rows 1-{max_row}, columns A-{get_column_letter(max_col)}")
print("="*100)
print("\n")

# Print every cell value
print("COMPLETE CELL DUMP:")
print("="*100)

for row_num in range(1, max_row + 1):
    row_has_data = False
    row_content = []

    for col_num in range(1, max_col + 1):
        cell = ws.cell(row=row_num, column=col_num)
        value = cell.value

        if value is not None and value != "":
            row_has_data = True
            col_letter = get_column_letter(col_num)
            row_content.append(f"  [{col_letter}]: {value}")

    # Print the row if it has any data
    if row_has_data:
        print(f"\nROW {row_num}:")
        for content in row_content:
            print(content)

print("\n")
print("="*100)
print("\nANALYSIS - Looking for key sections:")
print("="*100)

# Try to identify sections by looking for key patterns
sections = []
current_section = None

for row_num in range(1, max_row + 1):
    # Check first few columns for section headers or labels
    col_a = ws.cell(row=row_num, column=1).value
    col_b = ws.cell(row=row_num, column=2).value
    col_c = ws.cell(row=row_num, column=3).value

    # Look for patterns that indicate section starts
    if col_a and isinstance(col_a, str):
        col_a_lower = str(col_a).lower()

        # Check if this looks like a section header
        if any(keyword in col_a_lower for keyword in ['forecast', 'actual', 'budget', 'variance', '%', 'total', 'segment']):
            if current_section:
                sections.append(current_section)
            current_section = {
                'start_row': row_num,
                'header': col_a,
                'context': [col_a, col_b, col_c]
            }
        elif current_section:
            current_section['end_row'] = row_num

    # Check if row is completely empty (section break)
    row_empty = True
    for col_num in range(1, min(20, max_col + 1)):
        if ws.cell(row=row_num, column=col_num).value is not None:
            row_empty = False
            break

    if row_empty and current_section and 'end_row' not in current_section:
        current_section['end_row'] = row_num - 1
        sections.append(current_section)
        current_section = None

# Add last section if exists
if current_section:
    current_section['end_row'] = max_row
    sections.append(current_section)

print(f"\nIdentified {len(sections)} potential sections:\n")
for i, section in enumerate(sections, 1):
    print(f"Section {i}:")
    print(f"  Rows: {section['start_row']} to {section.get('end_row', 'end')}")
    print(f"  Header: {section['header']}")
    print(f"  Context: {section['context']}")
    print()

print("="*100)
print("\nSample from first 3 sections (first 5 columns, first 5 rows of each):")
print("="*100)

for i, section in enumerate(sections[:3], 1):
    print(f"\n--- Section {i} (starting at row {section['start_row']}) ---")
    start = section['start_row']
    end = min(start + 4, section.get('end_row', start + 4))

    for row_num in range(start, end + 1):
        row_data = []
        for col_num in range(1, 6):
            cell_value = ws.cell(row=row_num, column=col_num).value
            col_letter = get_column_letter(col_num)
            row_data.append(f"{col_letter}:{cell_value}")
        print(f"  Row {row_num}: {' | '.join(row_data)}")

print("\n" + "="*100)
print("Script complete.")

wb.close()
