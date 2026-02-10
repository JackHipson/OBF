"""
Create a complete map of the 'forecast vs actuals' sheet structure.
"""

import openpyxl
from openpyxl.utils import get_column_letter

file_path = "/home/user/OBF/Forecast Baseline Outputs v3.6 (new collections & impairment).xlsx"

# Open workbook
wb = openpyxl.load_workbook(file_path, data_only=True)
ws = wb["forecast vs actuals"]

print("="*100)
print("COMPLETE SHEET STRUCTURE MAP")
print("="*100)

sections = [
    {"start": 1, "end": 48, "name": "BB Forecast (detailed by metric x segment)"},
    {"start": 49, "end": 193, "name": "BB Forecast Summary & %s"},
    {"start": 194, "end": 368, "name": "ACTUALS (total book)"},
    {"start": 369, "end": ws.max_row, "name": "VARIANCES (forecast minus actuals)"}
]

for section in sections:
    print(f"\n{'='*100}")
    print(f"SECTION: {section['name']}")
    print(f"Rows {section['start']} to {section['end']}")
    print(f"{'='*100}")

    # Show first 15 rows of each section
    print(f"\nFirst 15 rows:")
    for row_num in range(section['start'], min(section['start'] + 15, section['end'] + 1)):
        row_data = []

        # Check key columns for labels
        for col_num in [5, 6, 15, 16]:  # E, F, O, P
            val = ws.cell(row=row_num, column=col_num).value
            if val is not None and val != "":
                col_letter = get_column_letter(col_num)
                val_str = str(val)
                if len(val_str) > 50:
                    val_str = val_str[:47] + "..."
                row_data.append(f"{col_letter}:{val_str}")

        # Check first data column (H for actuals, Q for forecast)
        for col_num in [8, 17]:
            val = ws.cell(row=row_num, column=col_num).value
            if val is not None and val != "" and val != 0:
                col_letter = get_column_letter(col_num)
                row_data.append(f"{col_letter}:{val}")
                break  # Only show one data sample

        if row_data:
            print(f"  Row {row_num}: {' | '.join(row_data)}")

# Now let's identify the specific metrics in each section
print(f"\n\n{'='*100}")
print("DETAILED METRICS IN EACH SECTION")
print(f"{'='*100}")

print("\n--- FORECAST SECTION (Rows 1-193) ---")
print("\nMetrics (Column O):")
for row_num in range(1, 194):
    val_o = ws.cell(row=row_num, column=15).value  # Column O
    if val_o and isinstance(val_o, str) and 'Sum of' in val_o:
        val_p = ws.cell(row=row_num, column=16).value
        print(f"  Row {row_num}: {val_o} (Segment: {val_p})")

print("\n\n--- ACTUALS SECTION (Rows 194-368) ---")
print("\nKey metrics (Column F):")
metrics_actuals = set()
for row_num in range(194, 369):
    val_f = ws.cell(row=row_num, column=6).value  # Column F
    if val_f and isinstance(val_f, str) and val_f.strip() != "":
        # Check if this is a metric label (not a segment)
        if not any(seg in val_f for seg in ['Non Prime', 'Near Prime', 'Prime']):
            metrics_actuals.add((row_num, val_f))

for row, metric in sorted(metrics_actuals):
    print(f"  Row {row}: {metric}")

print("\n\n--- VARIANCES SECTION (Rows 369+) ---")
print("\nSample rows (first 20):")
for row_num in range(369, min(389, ws.max_row + 1)):
    row_data = []
    for col_num in [5, 6, 15, 16]:  # E, F, O, P
        val = ws.cell(row=row_num, column=col_num).value
        if val is not None and val != "":
            col_letter = get_column_letter(col_num)
            val_str = str(val)
            if len(val_str) > 50:
                val_str = val_str[:47] + "..."
            row_data.append(f"{col_letter}:{val_str}")

    if row_data:
        print(f"  Row {row_num}: {' | '.join(row_data)}")

# Check month headers
print(f"\n\n{'='*100}")
print("MONTH HEADERS")
print(f"{'='*100}")

print("\nRow 1 (main headers):")
for col_num in range(8, 25):  # H to X
    val = ws.cell(row=1, column=col_num).value
    if val:
        print(f"  {get_column_letter(col_num)}: {val}")

print("\nRow 4 (forecast headers):")
for col_num in range(15, 25):  # O to X
    val = ws.cell(row=4, column=col_num).value
    if val:
        print(f"  {get_column_letter(col_num)}: {val}")

wb.close()
print("\n" + "="*100)
print("Complete structure map created")
print("="*100)
