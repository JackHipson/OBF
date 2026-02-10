"""
Extract key metrics from the 'forecast vs actuals' sheet.
Demonstrates how to pull specific metrics by row/column.

Usage:
    python extract_key_metrics.py
"""

import openpyxl
from openpyxl.utils import get_column_letter
import pandas as pd
from datetime import datetime

# File path
FILE_PATH = "/home/user/OBF/Forecast Baseline Outputs v3.6 (new collections & impairment).xlsx"
SHEET_NAME = "forecast vs actuals"

def get_month_from_column(ws, col_num):
    """Get the month/date from row 1 for a given column number."""
    date_val = ws.cell(row=1, column=col_num).value
    if date_val and isinstance(date_val, datetime):
        return date_val.strftime("%Y-%m")
    return None

def extract_metric_timeseries(ws, metric_row, start_col=8, end_col=17):
    """
    Extract a time series for a specific metric.

    Args:
        ws: Worksheet object
        metric_row: Row number of the metric
        start_col: Start column (default 8 = Column H = Jan 2025)
        end_col: End column (default 17 = Column Q = Oct 2025)

    Returns:
        Dictionary with {month: value}
    """
    data = {}
    for col_num in range(start_col, end_col + 1):
        month = get_month_from_column(ws, col_num)
        value = ws.cell(row=metric_row, column=col_num).value
        if month:
            data[month] = value
    return data

def main():
    print("="*100)
    print("EXTRACTING KEY METRICS FROM FORECAST VS ACTUALS SHEET")
    print("="*100)

    # Open workbook
    print(f"\nOpening workbook: {FILE_PATH}")
    wb = openpyxl.load_workbook(FILE_PATH, data_only=True)
    ws = wb[SHEET_NAME]
    print(f"Loaded sheet: {SHEET_NAME}")

    # Define key metrics
    key_metrics = {
        "Total Collections (P+I)": 200,
        "Total Collections inc Contra": 222,
        "Closing GBV": 240,
        "Average GBV": 250,
        "Closing NBV": 260,
        "Average NBV": 270,
        "Revenue": 280,
        "Gross Impairment": 291,
        "Net Impairment": 339,
        "RAM (incl DS gain)": 355
    }

    # Extract actuals data (Columns H-P = Jan-Sep 2025)
    print("\n" + "="*100)
    print("ACTUALS DATA (Jan 2025 - Sep 2025)")
    print("="*100)

    actuals_data = {}

    for metric_name, row_num in key_metrics.items():
        print(f"\n{metric_name} (Row {row_num}):")
        timeseries = extract_metric_timeseries(ws, row_num, start_col=8, end_col=16)
        actuals_data[metric_name] = timeseries

        for month, value in timeseries.items():
            if value is not None and value != 0:
                print(f"  {month}: {value:,.2f}")

    # Create a DataFrame for easier analysis
    print("\n" + "="*100)
    print("SUMMARY TABLE - ALL KEY METRICS")
    print("="*100)

    df = pd.DataFrame(actuals_data).T
    print("\n", df.to_string())

    # Export to CSV
    csv_path = "/home/user/OBF/actuals_key_metrics.csv"
    df.to_csv(csv_path)
    print(f"\n✓ Data exported to: {csv_path}")

    # Example: Calculate specific derived metrics
    print("\n" + "="*100)
    print("EXAMPLE CALCULATIONS")
    print("="*100)

    # Calculate Collections as % of Average GBV for Sep 2025
    sep_collections = ws['P200'].value  # Row 200, Column P
    sep_avg_gbv = ws['P250'].value      # Row 250, Column P

    if sep_collections and sep_avg_gbv and sep_avg_gbv != 0:
        coll_pct = (sep_collections / sep_avg_gbv) * 100
        print(f"\nSep 2025 Collections as % of Avg GBV: {coll_pct:.2f}%")

    # Get segment breakdown for Sep 2025 collections
    print("\nSep 2025 Collections by Segment:")
    segments = {
        "Non Prime": ws['P196'].value,
        "Near Prime Small": ws['P197'].value,
        "Near Prime Medium": ws['P198'].value,
        "Prime": ws['P199'].value
    }

    for segment, value in segments.items():
        if value:
            print(f"  {segment}: {value:,.2f}")

    # Example: Compare forecast to actuals for Oct 2025
    print("\n" + "="*100)
    print("FORECAST DATA SAMPLE (Oct 2025)")
    print("="*100)

    print("\nBB Forecast - Opening GBV by Segment (Oct 2025, Column Q):")
    forecast_segments = {
        "NON PRIME": ws['Q5'].value,
        "NRP-L": ws['Q6'].value,
        "NRP-M": ws['Q7'].value,
        "NRP-S": ws['Q8'].value,
        "PRIME": ws['Q9'].value
    }

    for segment, value in forecast_segments.items():
        if value:
            print(f"  {segment}: {value:,.2f}")

    print("\n" + "="*100)
    print("EXTRACTION COMPLETE")
    print("="*100)

    wb.close()

if __name__ == "__main__":
    main()
