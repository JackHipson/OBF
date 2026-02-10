"""
Analyze the structure of the 'forecast vs actuals' sheet to identify sections.
"""

import openpyxl
from openpyxl.utils import get_column_letter

file_path = "/home/user/OBF/Forecast Baseline Outputs v3.6 (new collections & impairment).xlsx"

# Open workbook
wb = openpyxl.load_workbook(file_path, data_only=True)
ws = wb["forecast vs actuals"]

print("="*100)
print("ANALYZING SHEET STRUCTURE")
print("="*100)

# Check dimensions
print(f"\nSheet dimensions: {ws.max_row} rows x {ws.max_column} columns")

# Look for key rows that might indicate section boundaries
print("\n" + "="*100)
print("SCANNING FOR SECTION MARKERS (checking column A-F for labels)")
print("="*100)

section_markers = []

for row_num in range(1, min(200, ws.max_row + 1)):
    # Check first 6 columns for section markers
    row_values = []
    for col_num in range(1, 7):
        val = ws.cell(row=row_num, column=col_num).value
        if val:
            row_values.append(f"{get_column_letter(col_num)}:{val}")

    # Check if this row might be a section marker
    col_a = ws.cell(row=row_num, column=1).value
    col_e = ws.cell(row=row_num, column=5).value
    col_f = ws.cell(row=row_num, column=6).value

    # Look for rows that look like headers or section starts
    if col_a and isinstance(col_a, str):
        col_a_str = str(col_a).upper()
        if any(keyword in col_a_str for keyword in ['FORECAST', 'ACTUAL', 'BUDGET', 'TOTAL', 'BOOK', '>>>', '<<<']):
            print(f"\nRow {row_num}: {' | '.join(row_values)}")
            section_markers.append(row_num)
    elif col_e and isinstance(col_e, str):
        col_e_str = str(col_e).upper()
        if any(keyword in col_e_str for keyword in ['FORECAST', 'ACTUAL', 'BUDGET', 'TOTAL', 'BOOK', '>>>', '<<<']):
            print(f"\nRow {row_num}: {' | '.join(row_values)}")
            section_markers.append(row_num)
    elif col_f and isinstance(col_f, str):
        col_f_str = str(col_f).upper()
        if any(keyword in col_f_str for keyword in ['FORECAST', 'ACTUAL', 'BUDGET', 'TOTAL', 'BOOK', '>>>', '<<<']):
            print(f"\nRow {row_num}: {' | '.join(row_values)}")
            section_markers.append(row_num)

print("\n" + "="*100)
print("DETAILED VIEW OF KEY ROWS")
print("="*100)

# Show detailed view of rows 45-70 (likely transition area)
print("\n--- ROWS 45-70 (Transition Area) ---")
for row_num in range(45, 71):
    row_data = []
    for col_num in range(1, 10):  # Columns A-I
        val = ws.cell(row=row_num, column=col_num).value
        if val is not None and val != "":
            col_letter = get_column_letter(col_num)
            row_data.append(f"{col_letter}:{val}")

    if row_data:
        print(f"Row {row_num}: {' | '.join(row_data)}")

# Check what's in column O (seems to contain metric names based on earlier output)
print("\n" + "="*100)
print("COLUMN O CONTENTS (appears to be metric names)")
print("="*100)

for row_num in range(1, min(100, ws.max_row + 1)):
    val_o = ws.cell(row=row_num, column=15).value  # Column O
    val_p = ws.cell(row=row_num, column=16).value  # Column P

    if val_o and str(val_o).strip() != "":
        print(f"Row {row_num}: O={val_o} | P={val_p}")

# Check column BF (last column in our scan, might be summary)
print("\n" + "="*100)
print("COLUMN BF CONTENTS (column 58 - might be summary)")
print("="*100)

for row_num in range(1, min(100, ws.max_row + 1)):
    val_bf = ws.cell(row=row_num, column=58).value  # Column BF
    val_f = ws.cell(row=row_num, column=6).value   # Column F for context

    if val_bf and val_bf != 0:
        print(f"Row {row_num}: F={val_f} | BF={val_bf}")

# Look for specific patterns indicating forecast vs actual sections
print("\n" + "="*100)
print("SEARCHING FOR PERCENTAGE OR VARIANCE INDICATORS")
print("="*100)

for row_num in range(1, min(150, ws.max_row + 1)):
    # Check if any cell in the row contains '%' or 'variance' or similar
    for col_num in range(1, 10):
        val = ws.cell(row=row_num, column=col_num).value
        if val and isinstance(val, str):
            val_lower = val.lower()
            if '%' in val_lower or 'variance' in val_lower or 'vs' in val_lower or 'diff' in val_lower:
                print(f"Row {row_num}, Col {get_column_letter(col_num)}: {val}")

wb.close()
print("\n" + "="*100)
print("Analysis complete")
print("="*100)
