"""
Examine the ACTUALS section starting at row 194.
"""

import openpyxl
from openpyxl.utils import get_column_letter

file_path = "/home/user/OBF/Forecast Baseline Outputs v3.6 (new collections & impairment).xlsx"

# Open workbook
wb = openpyxl.load_workbook(file_path, data_only=True)
ws = wb["forecast vs actuals"]

print("="*100)
print("ACTUALS SECTION - Rows 190-250")
print("="*100)

# Display rows 190-250 to see the actuals section structure
for row_num in range(190, 251):
    row_data = []

    # Check columns A-J for labels/headers
    for col_num in range(1, 11):
        val = ws.cell(row=row_num, column=col_num).value
        if val is not None and val != "":
            col_letter = get_column_letter(col_num)
            # Truncate long values
            val_str = str(val)
            if len(val_str) > 40:
                val_str = val_str[:37] + "..."
            row_data.append(f"{col_letter}:{val_str}")

    # Also check columns O, P for metric/segment names
    for col_num in [15, 16]:  # O and P
        val = ws.cell(row=row_num, column=col_num).value
        if val is not None and val != "":
            col_letter = get_column_letter(col_num)
            val_str = str(val)
            if len(val_str) > 40:
                val_str = val_str[:37] + "..."
            row_data.append(f"{col_letter}:{val_str}")

    # Check column Q (first data column in forecast section)
    val_q = ws.cell(row=row_num, column=17).value
    if val_q is not None and val_q != "":
        row_data.append(f"Q:{val_q}")

    # Check column BF (summary column)
    val_bf = ws.cell(row=row_num, column=58).value
    if val_bf is not None and val_bf != "":
        row_data.append(f"BF:{val_bf}")

    if row_data:
        print(f"\nRow {row_num}: {' | '.join(row_data)}")

print("\n" + "="*100)
print("CHECKING FOR VARIANCE/PERCENTAGE SECTIONS (Rows 250-350)")
print("="*100)

# Look for variance or percentage comparison rows
for row_num in range(250, 351):
    row_data = []

    # Check first 10 columns
    for col_num in range(1, 11):
        val = ws.cell(row=row_num, column=col_num).value
        if val is not None and val != "":
            col_letter = get_column_letter(col_num)
            val_str = str(val)
            if len(val_str) > 40:
                val_str = val_str[:37] + "..."
            row_data.append(f"{col_letter}:{val_str}")

    if row_data:
        # Check if it looks like a section header
        first_val = ws.cell(row=row_num, column=1).value
        if first_val:
            first_val_str = str(first_val).upper()
            if any(keyword in first_val_str for keyword in ['>>>', '<<<', 'VARIANCE', 'VS', 'COMPARISON', '%']):
                print(f"\n**Row {row_num}**: {' | '.join(row_data)}")
            else:
                # Just print rows with content
                col_f = ws.cell(row=row_num, column=6).value
                if col_f and any(keyword in str(col_f).lower() for keyword in ['%', 'vs', 'variance', 'diff']):
                    print(f"\nRow {row_num}: {' | '.join(row_data)}")

print("\n" + "="*100)
print("SUMMARY OF KEY SECTION BOUNDARIES")
print("="*100)

# Find all section markers
section_markers = []
for row_num in range(1, min(500, ws.max_row + 1)):
    for col_num in range(1, 10):
        val = ws.cell(row=row_num, column=col_num).value
        if val and isinstance(val, str):
            if '>>>' in val or '<<<' in val:
                section_markers.append({
                    'row': row_num,
                    'col': get_column_letter(col_num),
                    'text': val
                })
                break

print("\nSection markers found:")
for marker in section_markers:
    print(f"  Row {marker['row']} (Col {marker['col']}): {marker['text']}")

wb.close()
print("\n" + "="*100)
print("Analysis complete")
print("="*100)
